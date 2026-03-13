"""Excel file parser using openpyxl."""

import uuid
from pathlib import Path

from openpyxl import load_workbook

from app.ingestion.parsers.base import BaseParser
from app.models.internal import Document


class ExcelParser(BaseParser):
    """Parser for Excel files (.xlsx, .xls) using openpyxl."""

    def parse(self, file_path: Path) -> Document:
        """Parse an Excel file."""
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        stat = file_path.stat()
        content = self._extract_text(file_path)

        return Document(
            doc_id=str(uuid.uuid4())[:8],
            file_path=str(file_path),
            content=content,
            file_type=file_path.suffix,
            file_size=stat.st_size,
        )

    def _extract_text(self, file_path: Path) -> str:
        """Extract text from Excel workbook."""
        wb = load_workbook(str(file_path), data_only=True, read_only=True)
        text_parts: list[str] = []

        try:
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_content = self._extract_sheet(sheet)
                if sheet_content.strip():
                    text_parts.append(f"[Sheet: {sheet_name}]\n{sheet_content}")
        finally:
            wb.close()

        return "\n\n".join(text_parts)

    def _extract_sheet(self, sheet) -> str:  # type: ignore
        """Extract text from a single worksheet."""
        rows: list[str] = []

        for row in sheet.iter_rows():
            cell_values = []
            for cell in row:
                if cell.value is not None:
                    cell_values.append(str(cell.value))
            if cell_values:
                rows.append(" | ".join(cell_values))

        return "\n".join(rows)

    @classmethod
    def supported_extensions(cls) -> list[str]:
        return [".xlsx", ".xls"]
